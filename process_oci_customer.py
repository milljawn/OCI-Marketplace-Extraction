#!/usr/bin/env python3
"""
OCI Marketplace Data Processor - Customer Configuration (Fixed)
Processes Commercial (OC1), DoD (OC2), and US Gov (OC3) marketplace data into comprehensive Excel format
Compatible with updated shell script output files
"""

import json
import pandas as pd
import os
from datetime import datetime
import re
from collections import defaultdict

class OCIMarketplaceSalesProcessor:
    def __init__(self, data_dir="marketplace_data"):
        self.data_dir = data_dir
        self.all_listings = {}  # Dict of region -> listings
        self.unique_listings = {}  # Dict of listing_id -> consolidated listing info
        
        # Updated file mappings based on the new shell script
        self.region_files = [
            ("commercial", "all_listings_commercial.json", "Commercial (OC1)"),
            ("oc3_us_gov_east", "oc3_us_gov_east_listings.json", "OC3 US Gov East"),
            ("oc3_us_gov_west", "oc3_us_gov_west_listings.json", "OC3 US Gov West"),
            ("oc2_us_dod_east", "oc2_us_dod_east_listings.json", "OC2 DoD East (Langley)"),
            ("oc2_us_dod_west", "oc2_us_dod_west_listings.json", "OC2 DoD West (Luke)"),
            ("legacy_us_dod_east", "legacy_us_dod_east_listings.json", "Legacy DoD East"),
            ("legacy_us_dod_central", "legacy_us_dod_central_listings.json", "Legacy DoD Central"),
            ("legacy_us_dod_west", "legacy_us_dod_west_listings.json", "Legacy DoD West")
        ]
        
    def safe_load_json(self, filepath):
        """Safely load JSON file with error handling"""
        if not os.path.exists(filepath):
            return {}
        
        if os.path.getsize(filepath) == 0:
            return {}
        
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                content = f.read().strip()
                if not content:
                    return {}
                
                data = json.loads(content)
                return data
                
        except json.JSONDecodeError as e:
            print(f"❌ JSON decode error in {filepath}: {e}")
            return {}
        except Exception as e:
            print(f"❌ Error loading {filepath}: {e}")
            return {}
    
    def load_all_regional_data(self):
        """Load marketplace data from all regions"""
        print("Loading marketplace data from customer's multi-realm OCI setup...")
        print("=" * 65)
        
        total_listings = 0
        regions_found = []
        
        for region_key, filename, region_name in self.region_files:
            filepath = os.path.join(self.data_dir, filename)
            data = self.safe_load_json(filepath)
            
            if data and 'data' in data and len(data['data']) > 0:
                self.all_listings[region_key] = data['data']
                count = len(data['data'])
                total_listings += count
                regions_found.append(region_name)
                print(f"✅ {region_name:<25}: {count:>4} listings")
                
                # Load detailed data if available
                detailed_file = filepath.replace("_listings.json", "_detailed.json")
                detailed_data = self.safe_load_json(detailed_file)
                if detailed_data and 'data' in detailed_data and len(detailed_data['data']) > 0:
                    self.merge_detailed_data(region_key, detailed_data['data'])
                    print(f"   {'└─ Detailed metadata merged':<25}")
            else:
                self.all_listings[region_key] = []
                if os.path.exists(filepath):
                    print(f"⚠️  {region_name:<25}: File exists but empty")
                else:
                    print(f"   {region_name:<25}: Not extracted (no access)")
        
        print(f"\n📊 Total regional listings loaded: {total_listings}")
        print(f"🌍 Active regions: {len(regions_found)}")
        if regions_found:
            print(f"   ✓ {', '.join(regions_found)}")
        
        return total_listings
    
    def merge_detailed_data(self, region_key, detailed_data):
        """Merge detailed data with main listings"""
        if not detailed_data or region_key not in self.all_listings:
            return
            
        detailed_dict = {item.get('id'): item for item in detailed_data if item.get('id')}
        
        for listing in self.all_listings[region_key]:
            listing_id = listing.get('id')
            if listing_id in detailed_dict:
                detailed_item = detailed_dict[listing_id]
                for key, value in detailed_item.items():
                    if key not in listing or listing[key] is None:
                        listing[key] = value
    
    def consolidate_unique_listings(self):
        """Consolidate listings across regions to identify unique products"""
        print("\nConsolidating unique listings across all realms...")
        
        region_mapping = {
            'commercial': 'Commercial (OC1)',
            'oc3_us_gov_east': 'US Gov East (OC3)',
            'oc3_us_gov_west': 'US Gov West (OC3)',
            'oc2_us_dod_east': 'DoD East/Langley (OC2)',
            'oc2_us_dod_west': 'DoD West/Luke (OC2)',
            'legacy_us_dod_east': 'Legacy DoD East',
            'legacy_us_dod_central': 'Legacy DoD Central',
            'legacy_us_dod_west': 'Legacy DoD West'
        }
        
        for region_key, listings in self.all_listings.items():
            for listing in listings:
                listing_id = listing.get('id', '')
                
                if listing_id not in self.unique_listings:
                    # First time seeing this listing
                    self.unique_listings[listing_id] = {
                        'listing': listing,
                        'regions': {region_key: True},
                        'primary_region': region_key,
                        'region_names': [region_mapping.get(region_key, region_key)]
                    }
                else:
                    # Listing exists in multiple regions
                    self.unique_listings[listing_id]['regions'][region_key] = True
                    region_name = region_mapping.get(region_key, region_key)
                    if region_name not in self.unique_listings[listing_id]['region_names']:
                        self.unique_listings[listing_id]['region_names'].append(region_name)
                    
                    # Merge any additional data from this region's version
                    existing = self.unique_listings[listing_id]['listing']
                    for key, value in listing.items():
                        if key not in existing or existing[key] is None or existing[key] == '':
                            existing[key] = value
        
        print(f"✅ Identified {len(self.unique_listings)} unique marketplace listings")
        
        # Print region overlap statistics
        region_counts = defaultdict(int)
        for listing_info in self.unique_listings.values():
            num_regions = len(listing_info['regions'])
            region_counts[num_regions] += 1
        
        print("\nRegion availability distribution:")
        for num_regions, count in sorted(region_counts.items()):
            regions_text = "region" if num_regions == 1 else "regions"
            print(f"   Available in {num_regions} {regions_text}: {count:>4} products")
        
        return len(self.unique_listings)
    
    def process_to_sales_excel(self):
        """Convert consolidated data to sales-focused Excel format with multiple sheets"""
        print("\nProcessing data for global sales organization Excel format...")
        
        # Main catalog data
        catalog_data = []
        
        # Analytics data for summary sheets
        publisher_stats = defaultdict(lambda: {'total': 0, 'gov': 0, 'dod': 0, 'commercial': 0})
        category_stats = defaultdict(lambda: {'total': 0, 'gov': 0, 'dod': 0, 'commercial': 0})
        region_stats = defaultdict(int)
        
        for listing_id, listing_info in self.unique_listings.items():
            listing = listing_info['listing']
            regions = listing_info['regions']
            
            try:
                # Core product information
                product_name = self.safe_get(listing, 'name')
                publisher_name = self.get_publisher_name(listing)
                category = self.get_category(listing)
                
                # Determine availability flags
                commercial_available = bool(regions.get('commercial'))
                us_gov_available = bool(regions.get('oc3_us_gov_east') or regions.get('oc3_us_gov_west'))
                dod_available = bool(regions.get('oc2_us_dod_east') or regions.get('oc2_us_dod_west') or 
                                   any(k.startswith('legacy_us_dod') for k in regions))
                
                # Government authorization assessment
                gov_auth = self.determine_gov_authorization_level(regions)
                
                # Sales intelligence
                sales_priority = self.calculate_sales_priority_score(listing, regions)
                gov_sales_priority = self.calculate_gov_sales_priority(listing, regions)
                
                # Create main catalog row
                row = {
                    # Core Product Info
                    'Product_Name': product_name,
                    'Publisher': publisher_name,
                    'Category': category,
                    'Short_Description': self.clean_text(self.safe_get(listing, 'short-description'), 150),
                    'Package_Type': self.safe_get(listing, 'package-type', 'Standard'),
                    
                    # Availability Matrix
                    'Commercial_OC1': 'Yes' if commercial_available else 'No',
                    'US_Gov_OC3': 'Yes' if us_gov_available else 'No',
                    'DoD_OC2': 'Yes' if dod_available else 'No',
                    'Total_Regions': len(listing_info['regions']),
                    'Available_Regions': ' | '.join(listing_info['region_names']),
                    
                    # Government & Compliance
                    'Gov_Authorization_Level': gov_auth['level'],
                    'Gov_Clearance_Details': gov_auth['details'],
                    'FedRAMP_Status': self.determine_fedramp_status(listing),
                    'DoD_Impact_Level': self.determine_impact_level(listing),
                    'CMMC_Level': self.determine_cmmc_level(listing),
                    'Security_Certifications': self.get_security_certifications(listing),
                    
                    # Pricing & Business Model
                    'Pricing_Model': self.get_pricing_model(listing),
                    'Estimated_Price': self.get_estimated_price(listing),
                    'Free_Trial_Available': 'Yes' if self.safe_get(listing, 'free-trial-available') else 'No',
                    'Oracle_Validated': 'Yes' if self.safe_get(listing, 'oracle-validated') else 'No',
                    
                    # Sales Intelligence
                    'Sales_Priority_Score': sales_priority,
                    'Gov_Sales_Priority': gov_sales_priority,
                    'Target_Customer_Size': self.determine_target_customer_size(listing),
                    'Industry_Focus': self.determine_industry_focus(listing),
                    'Primary_Use_Case': self.extract_primary_use_case(listing),
                    
                    # Technical Details
                    'Deployment_Method': self.get_deployment_method(listing),
                    'Integration_Complexity': self.assess_integration_complexity(listing),
                    'Supported_OS': self.get_supported_platforms(listing),
                    'Support_Type': self.determine_support_type(listing),
                    
                    # Sales Enablement
                    'Demo_Available': self.check_demo_availability(listing),
                    'POC_Duration_Est': self.estimate_poc_duration(listing),
                    'Implementation_Timeline': self.estimate_implementation_timeline(listing),
                    'Reference_Customers': self.check_reference_availability(listing),
                    
                    # Competitive Intelligence
                    'Key_Competitors': self.identify_competitors(listing),
                    'Unique_Value_Prop': self.extract_value_proposition(listing),
                    'Key_Features': self.extract_key_features(listing),
                    
                    # URLs & Resources
                    'Marketplace_URL': self.generate_marketplace_url(listing_id),
                    'Documentation_URL': self.safe_get(listing, 'documentation-url'),
                    'Support_URL': self.safe_get(listing, 'support-url'),
                    'Video_Demo_URL': self.safe_get(listing, 'video-url'),
                    
                    # Sales Notes & Strategy
                    'Gov_Sales_Notes': self.generate_gov_sales_notes(listing, regions),
                    'Sales_Strategy_Notes': self.generate_sales_strategy_notes(listing, regions),
                    
                    # Metadata
                    'Listing_ID': listing_id.replace('ocid1.marketplace.listing.', ''),
                    'Last_Updated': self.format_date(self.safe_get(listing, 'time-updated')),
                    'Oracle_Partner_Level': self.determine_partner_level(listing),
                    'Data_Extraction_Date': datetime.now().strftime('%Y-%m-%d')
                }
                
                catalog_data.append(row)
                
                # Update analytics
                publisher_stats[publisher_name]['total'] += 1
                category_stats[category]['total'] += 1
                
                if commercial_available:
                    publisher_stats[publisher_name]['commercial'] += 1
                    category_stats[category]['commercial'] += 1
                if us_gov_available:
                    publisher_stats[publisher_name]['gov'] += 1
                    category_stats[category]['gov'] += 1
                if dod_available:
                    publisher_stats[publisher_name]['dod'] += 1
                    category_stats[category]['dod'] += 1
                
                for region_name in listing_info['region_names']:
                    region_stats[region_name] += 1
                
            except Exception as e:
                print(f"⚠️  Error processing listing {listing_id}: {e}")
                continue
        
        print(f"✅ Processed {len(catalog_data)} listings for sales organization")
        
        return catalog_data, publisher_stats, category_stats, region_stats
    
    def create_government_analysis_sheet(self, catalog_df):
        """Create specialized government analysis sheet"""
        print("Creating government opportunities analysis...")
        
        # Filter for government-available products
        gov_products = catalog_df[
            (catalog_df['US_Gov_OC3'] == 'Yes') | 
            (catalog_df['DoD_OC2'] == 'Yes')
        ].copy()
        
        if len(gov_products) == 0:
            return pd.DataFrame()
        
        # Add government-specific analytics
        gov_products['Multi_Realm_Available'] = (
            (gov_products['US_Gov_OC3'] == 'Yes').astype(int) +
            (gov_products['DoD_OC2'] == 'Yes').astype(int)
        )
        
        gov_products['Compliance_Score'] = gov_products.apply(self.calculate_compliance_score, axis=1)
        gov_products['Gov_Market_Potential'] = gov_products.apply(self.assess_gov_market_potential, axis=1)
        
        # Select key columns for government analysis
        gov_analysis_columns = [
            'Product_Name', 'Publisher', 'Category',
            'US_Gov_OC3', 'DoD_OC2',
            'Gov_Authorization_Level', 'Gov_Sales_Priority',
            'FedRAMP_Status', 'DoD_Impact_Level', 'CMMC_Level',
            'Security_Certifications', 'Compliance_Score',
            'Multi_Realm_Available', 'Gov_Market_Potential',
            'Pricing_Model', 'Target_Customer_Size',
            'Gov_Sales_Notes', 'Sales_Strategy_Notes',
            'Marketplace_URL'
        ]
        
        return gov_products[gov_analysis_columns]
    
    def create_publisher_analysis_sheet(self, publisher_stats):
        """Create publisher performance analysis sheet"""
        print("Creating publisher ecosystem analysis...")
        
        publisher_data = []
        for publisher, stats in publisher_stats.items():
            if stats['total'] > 0:  # Only include publishers with products
                gov_coverage = ((stats['gov'] + stats['dod']) / stats['total']) * 100
                
                publisher_data.append({
                    'Publisher_Name': publisher,
                    'Total_Products': stats['total'],
                    'Commercial_Products': stats['commercial'],
                    'US_Gov_Products': stats['gov'],
                    'DoD_Products': stats['dod'],
                    'Gov_Coverage_Percent': round(gov_coverage, 1),
                    'Publisher_Tier': self.classify_publisher_tier(publisher, stats),
                    'Strategic_Value': self.assess_publisher_strategic_value(stats),
                    'Recommended_Action': self.recommend_publisher_action(publisher, stats)
                })
        
        return pd.DataFrame(publisher_data).sort_values('Total_Products', ascending=False)
    
    def create_executive_summary_sheet(self, catalog_df, region_stats):
        """Create executive summary dashboard"""
        print("Creating executive summary dashboard...")
        
        # Key metrics
        total_products = len(catalog_df)
        commercial_products = len(catalog_df[catalog_df['Commercial_OC1'] == 'Yes'])
        gov_products = len(catalog_df[catalog_df['US_Gov_OC3'] == 'Yes'])
        dod_products = len(catalog_df[catalog_df['DoD_OC2'] == 'Yes'])
        
        # Priority breakdown
        priority_counts = catalog_df['Gov_Sales_Priority'].value_counts()
        
        # Top categories
        top_categories = catalog_df['Category'].value_counts().head(10)
        
        # Create summary data
        summary_data = []
        
        # Overall metrics
        summary_data.extend([
            {'Metric': 'Total Marketplace Products', 'Value': total_products, 'Category': 'Overview'},
            {'Metric': 'Commercial Cloud (OC1)', 'Value': commercial_products, 'Category': 'Availability'},
            {'Metric': 'US Government Cloud (OC3)', 'Value': gov_products, 'Category': 'Availability'},
            {'Metric': 'DoD Cloud (OC2)', 'Value': dod_products, 'Category': 'Availability'},
        ])
        
        # Government priorities
        for priority, count in priority_counts.items():
            summary_data.append({
                'Metric': f'Gov Sales Priority - {priority}',
                'Value': count,
                'Category': 'Government Priorities'
            })
        
        # Top categories
        for i, (category, count) in enumerate(top_categories.items(), 1):
            summary_data.append({
                'Metric': f'Top Category #{i} - {category}',
                'Value': count,
                'Category': 'Product Categories'
            })
        
        # Compliance overview
        fedramp_products = len(catalog_df[~catalog_df['FedRAMP_Status'].str.contains('Not Specified')])
        cmmc_products = len(catalog_df[~catalog_df['CMMC_Level'].str.contains('Not Specified')])
        
        summary_data.extend([
            {'Metric': 'FedRAMP Compliant Products', 'Value': fedramp_products, 'Category': 'Compliance'},
            {'Metric': 'CMMC Ready Products', 'Value': cmmc_products, 'Category': 'Compliance'},
        ])
        
        return pd.DataFrame(summary_data)
    
    def export_to_excel(self, filename='oracle_marketplace_global_sales_catalog.xlsx'):
        """Export to comprehensive Excel file with multiple sheets"""
        print(f"\n🏢 Generating global sales organization Excel catalog...")
        
        try:
            catalog_data, publisher_stats, category_stats, region_stats = self.process_to_sales_excel()
            
            if not catalog_data:
                print("❌ No data to export")
                return
            
            # Create main DataFrame
            catalog_df = pd.DataFrame(catalog_data)
            
            # Sort by government sales priority and overall priority
            priority_order = {'CRITICAL': 4, 'HIGH': 3, 'MEDIUM': 2, 'LOW': 1}
            catalog_df['_Gov_Priority_Sort'] = catalog_df['Gov_Sales_Priority'].map(priority_order).fillna(0)
            catalog_df['_Sales_Priority_Sort'] = pd.to_numeric(catalog_df['Sales_Priority_Score'], errors='coerce').fillna(0)
            
            catalog_df = catalog_df.sort_values(['_Gov_Priority_Sort', '_Sales_Priority_Sort'], ascending=[False, False])
            catalog_df = catalog_df.drop(['_Gov_Priority_Sort', '_Sales_Priority_Sort'], axis=1)
            
            # Create additional analysis sheets
            gov_analysis_df = self.create_government_analysis_sheet(catalog_df)
            publisher_df = self.create_publisher_analysis_sheet(publisher_stats)
            executive_df = self.create_executive_summary_sheet(catalog_df, region_stats)
            
            # Export to Excel with multiple sheets
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                # Main catalog
                catalog_df.to_excel(writer, sheet_name='Complete Catalog', index=False)
                
                # Government analysis
                if not gov_analysis_df.empty:
                    gov_analysis_df.to_excel(writer, sheet_name='Government Opportunities', index=False)
                
                # Publisher analysis
                publisher_df.to_excel(writer, sheet_name='Publisher Analysis', index=False)
                
                # Executive summary
                executive_df.to_excel(writer, sheet_name='Executive Summary', index=False)
                
                # Format sheets
                self.format_excel_sheets(writer, catalog_df, gov_analysis_df, publisher_df, executive_df)
            
            print(f"✅ Global sales catalog exported: {filename}")
            self.print_export_summary(catalog_df, gov_analysis_df, publisher_df)
            
        except Exception as e:
            print(f"❌ Error creating Excel file: {e}")
            import traceback
            traceback.print_exc()
    
    def format_excel_sheets(self, writer, catalog_df, gov_df, publisher_df, executive_df):
        """Format Excel sheets for professional appearance"""
        try:
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            # Color scheme
            header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
            priority_high = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
            priority_medium = PatternFill(start_color="FFE66D", end_color="FFE66D", fill_type="solid")
            
            # Format main catalog sheet
            ws_catalog = writer.sheets['Complete Catalog']
            for cell in ws_catalog[1]:  # Header row
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Auto-adjust column widths
            for column in ws_catalog.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_catalog.column_dimensions[column_letter].width = adjusted_width
            
        except ImportError:
            print("⚠️  openpyxl styling not available - basic formatting applied")
        except Exception as e:
            print(f"⚠️  Formatting error: {e}")
    
    def print_export_summary(self, catalog_df, gov_df, publisher_df):
        """Print export summary statistics"""
        print(f"\n📊 Export Summary:")
        print(f"   📄 Complete Catalog: {len(catalog_df)} products")
        print(f"   🏛️ Government Products: {len(gov_df)} products")
        print(f"   🏢 Publishers Analyzed: {len(publisher_df)} companies")
        
        print(f"\n🌍 Regional Breakdown:")
        print(f"   Commercial (OC1): {len(catalog_df[catalog_df['Commercial_OC1'] == 'Yes'])}")
        print(f"   US Government (OC3): {len(catalog_df[catalog_df['US_Gov_OC3'] == 'Yes'])}")
        print(f"   DoD (OC2): {len(catalog_df[catalog_df['DoD_OC2'] == 'Yes'])}")
        print(f"   UK Government: {len(catalog_df[catalog_df['UK_Gov'] == 'Yes'])}")
        
        print(f"\n🎯 Government Sales Priorities:")
        priority_counts = catalog_df['Gov_Sales_Priority'].value_counts()
        for priority, count in priority_counts.items():
            print(f"   {priority}: {count} products")
    
    # Helper methods (consolidated and improved versions of existing methods)
    
    def safe_get(self, data, keys, default=''):
        """Safely extract nested data"""
        if isinstance(keys, str):
            return data.get(keys, default) or default
        
        current = data
        for key in keys:
            if isinstance(current, dict) and key in current:
                current = current[key]
            else:
                return default
        return current if current is not None else default
    
    def clean_text(self, text, max_length=None):
        """Clean text for Excel output"""
        if not text:
            return ''
        
        # Remove excessive whitespace and newlines
        text = ' '.join(str(text).split())
        
        # Remove special characters that might cause issues in Excel
        text = re.sub(r'[^\w\s\-\.\,\(\)\[\]\/\:\;]', '', text)
        
        # Truncate if needed
        if max_length and len(text) > max_length:
            text = text[:max_length-3] + '...'
        
        return text
    
    def get_publisher_name(self, listing):
        """Extract publisher name"""
        publisher = self.safe_get(listing, 'publisher', {})
        if isinstance(publisher, dict):
            name = publisher.get('name') or publisher.get('display-name')
            if name:
                return self.clean_text(name, 50)
        
        return self.clean_text(str(publisher), 50) if publisher else 'Unknown Publisher'
    
    def get_category(self, listing):
        """Get primary category"""
        categories = self.safe_get(listing, 'categories') or self.safe_get(listing, 'category-facet')
        if categories:
            if isinstance(categories, list) and len(categories) > 0:
                return self.clean_text(categories[0], 30)
            elif isinstance(categories, str):
                return self.clean_text(categories, 30)
        return 'Uncategorized'
    
    def determine_gov_authorization_level(self, regions):
        """Determine government authorization level"""
        if regions.get('oc2_us_dod_east') or regions.get('oc2_us_dod_west'):
            return {
                'level': 'DoD Impact Level Ready',
                'details': 'Authorized for DoD classified environments'
            }
        elif any(k.startswith('legacy_us_dod') for k in regions):
            return {
                'level': 'DoD Legacy Authorized',
                'details': 'Available in legacy DoD regions'
            }
        elif regions.get('oc3_us_gov_east') or regions.get('oc3_us_gov_west'):
            return {
                'level': 'FedRAMP Authority',
                'details': 'Authorized for US Government use'
            }
        else:
            return {
                'level': 'Commercial Only',
                'details': 'Not yet government authorized'
            }
    
    def calculate_sales_priority_score(self, listing, regions):
        """Calculate sales priority score (1-10)"""
        score = 5  # Base score
        
        # Government availability boost
        if regions.get('oc2_us_dod_east') or regions.get('oc2_us_dod_west'):
            score += 3
        elif any(k.startswith('oc3_us_gov') for k in regions):
            score += 2
        elif any(k.startswith('legacy_us_dod') for k in regions):
            score += 1
        
        # Publisher reputation
        publisher = self.get_publisher_name(listing).lower()
        tier1_publishers = ['oracle', 'microsoft', 'vmware', 'cisco', 'palo alto', 'fortinet']
        if any(pub in publisher for pub in tier1_publishers):
            score += 2
        
        # Category importance for government
        category = self.get_category(listing).lower()
        high_value_categories = ['security', 'networking', 'database', 'analytics', 'monitoring']
        if any(cat in category for cat in high_value_categories):
            score += 1
        
        return min(10, max(1, score))
    
    def calculate_gov_sales_priority(self, listing, regions):
        """Calculate government-specific sales priority"""
        score = 0
        
        # OC2 DoD availability is highest priority
        if regions.get('oc2_us_dod_east') or regions.get('oc2_us_dod_west'):
            score += 6
        elif any(k.startswith('legacy_us_dod') for k in regions):
            score += 4
        
        # OC3 Government availability
        if regions.get('oc3_us_gov_east') or regions.get('oc3_us_gov_west'):
            score += 3
        
        # Security focus bonus
        if self.is_security_focused(listing):
            score += 2
        
        # Compliance indicators
        if self.determine_fedramp_status(listing) != 'Not Specified':
            score += 1
        if self.determine_cmmc_level(listing) != 'Not Specified':
            score += 1
        
        if score >= 8:
            return 'CRITICAL'
        elif score >= 5:
            return 'HIGH'
        elif score >= 3:
            return 'MEDIUM'
        else:
            return 'LOW'
    
    def get_pricing_model(self, listing):
        """Determine pricing model"""
        pricing = self.safe_get(listing, ['pricing', 'type'])
        
        pricing_map = {
            'FREE': 'Free',
            'BYOL': 'Bring Your Own License',
            'PAID': 'Pay-As-You-Go',
            'SUBSCRIPTION': 'Subscription',
            'USAGE': 'Usage-Based'
        }
        
        return pricing_map.get(pricing, pricing or 'Contact Sales')
    
    def get_estimated_price(self, listing):
        """Get estimated price"""
        pricing = self.safe_get(listing, 'pricing', {})
        pricing_type = pricing.get('type')
        
        if pricing_type == 'FREE':
            return 'Free'
        elif pricing_type == 'BYOL':
            return 'License Required'
        
        rate = pricing.get('rate')
        if rate:
            unit = pricing.get('unit', 'hour')
            currency = pricing.get('currency', 'USD')
            return f"{rate} {currency}/{unit}"
        
        return 'Contact for Pricing'
    
    def determine_fedramp_status(self, listing):
        """Determine FedRAMP status"""
        text_content = self.get_listing_text_content(listing)
        
        if 'fedramp high' in text_content:
            return 'FedRAMP High'
        elif 'fedramp moderate' in text_content:
            return 'FedRAMP Moderate'
        elif 'fedramp low' in text_content or 'fedramp authorized' in text_content:
            return 'FedRAMP Low'
        elif 'fedramp' in text_content:
            return 'FedRAMP Ready'
        elif any(term in text_content for term in ['fisma', 'federal compliance']):
            return 'Federal Ready'
        
        return 'Not Specified'
    
    def determine_impact_level(self, listing):
        """Determine DoD Impact Level"""
        text_content = self.get_listing_text_content(listing)
        
        if 'il6' in text_content or 'impact level 6' in text_content:
            return 'IL6 (Secret)'
        elif 'il5' in text_content or 'impact level 5' in text_content:
            return 'IL5 (CUI High)'
        elif 'il4' in text_content or 'impact level 4' in text_content:
            return 'IL4 (CUI)'
        elif 'il2' in text_content or 'impact level 2' in text_content:
            return 'IL2 (Unclassified)'
        elif any(term in text_content for term in ['dod', 'defense', 'classified']):
            return 'DoD Compatible'
        
        return 'Not Specified'
    
    def determine_cmmc_level(self, listing):
        """Determine CMMC Level"""
        text_content = self.get_listing_text_content(listing)
        
        if 'cmmc level 3' in text_content or 'cmmc l3' in text_content:
            return 'CMMC Level 3'
        elif 'cmmc level 2' in text_content or 'cmmc l2' in text_content:
            return 'CMMC Level 2'
        elif 'cmmc' in text_content:
            return 'CMMC Level 1+'
        
        return 'Not Specified'
    
    def get_security_certifications(self, listing):
        """Extract security certifications"""
        text_content = self.get_listing_text_content(listing)
        
        certs = []
        cert_keywords = {
            'SOC 2': ['soc 2', 'soc2', 'soc ii'],
            'ISO 27001': ['iso 27001', 'iso27001'],
            'PCI DSS': ['pci dss', 'pci-dss'],
            'HIPAA': ['hipaa'],
            'NIST': ['nist 800-53', 'nist framework'],
            'FIPS 140-2': ['fips 140-2', 'fips 140']
        }
        
        for cert_name, keywords in cert_keywords.items():
            if any(keyword in text_content for keyword in keywords):
                certs.append(cert_name)
        
        return '; '.join(certs) if certs else 'Standard Compliance'
    
    def determine_target_customer_size(self, listing):
        """Determine target customer size"""
        text_content = self.get_listing_text_content(listing)
        
        if any(term in text_content for term in ['enterprise', 'large organization', 'scalable']):
            return 'Enterprise'
        elif any(term in text_content for term in ['small business', 'startup', 'basic']):
            return 'Small Business'
        else:
            return 'Mid-Market'
    
    def determine_industry_focus(self, listing):
        """Determine industry focus"""
        text_content = self.get_listing_text_content(listing)
        
        industries = []
        industry_map = {
            'Government': ['government', 'federal', 'public sector'],
            'Defense': ['defense', 'military', 'dod'],
            'Healthcare': ['healthcare', 'medical', 'hipaa'],
            'Financial': ['financial', 'banking', 'fintech'],
            'Retail': ['retail', 'ecommerce', 'commerce'],
            'Manufacturing': ['manufacturing', 'industrial'],
            'Education': ['education', 'academic', 'university']
        }
        
        for industry, keywords in industry_map.items():
            if any(keyword in text_content for keyword in keywords):
                industries.append(industry)
        
        return '; '.join(industries) if industries else 'Cross-Industry'
    
    def extract_primary_use_case(self, listing):
        """Extract primary use case"""
        category = self.get_category(listing).lower()
        
        use_case_map = {
            'security': 'Cybersecurity & Compliance',
            'database': 'Data Management',
            'analytics': 'Business Intelligence',
            'networking': 'Network Infrastructure',
            'compute': 'Cloud Computing',
            'storage': 'Data Storage',
            'monitoring': 'Performance Monitoring',
            'developer': 'Application Development'
        }
        
        for key, use_case in use_case_map.items():
            if key in category:
                return use_case
        
        return 'Enterprise Software'
    
    def get_deployment_method(self, listing):
        """Get deployment method"""
        package_type = str(self.safe_get(listing, 'package-type', '')).upper()
        
        deployment_map = {
            'IMAGE': 'Virtual Machine Image',
            'STACK': 'Oracle Stack',
            'TERRAFORM': 'Terraform Template',
            'KUBERNETES': 'Kubernetes Application',
            'CONTAINER': 'Container Image'
        }
        
        return deployment_map.get(package_type, 'Standard Deployment')
    
    def assess_integration_complexity(self, listing):
        """Assess integration complexity"""
        package_type = str(self.safe_get(listing, 'package-type', '')).upper()
        text_content = self.get_listing_text_content(listing)
        
        if package_type == 'IMAGE':
            return 'Low'
        elif any(term in text_content for term in ['complex', 'enterprise integration', 'custom']):
            return 'High'
        else:
            return 'Medium'
    
    def get_supported_platforms(self, listing):
        """Get supported platforms"""
        os_list = self.safe_get(listing, 'supported-operating-systems', [])
        if os_list and isinstance(os_list, list):
            return '; '.join(os_list[:3])  # Limit to first 3 for readability
        
        return 'Oracle Linux'
    
    def determine_support_type(self, listing):
        """Determine support type"""
        publisher = self.get_publisher_name(listing).lower()
        
        if 'oracle' in publisher:
            return 'Oracle Support'
        elif self.safe_get(listing, 'support-url'):
            return 'Vendor Support'
        else:
            return 'Community Support'
    
    def check_demo_availability(self, listing):
        """Check demo availability"""
        if self.safe_get(listing, 'video-url'):
            return 'Video Demo Available'
        elif self.safe_get(listing, 'free-trial-available'):
            return 'Free Trial Available'
        else:
            return 'Contact Vendor'
    
    def estimate_poc_duration(self, listing):
        """Estimate POC duration"""
        complexity = self.assess_integration_complexity(listing)
        
        if complexity == 'Low':
            return '1-2 weeks'
        elif complexity == 'Medium':
            return '2-4 weeks'
        else:
            return '4-8 weeks'
    
    def estimate_implementation_timeline(self, listing):
        """Estimate implementation timeline"""
        complexity = self.assess_integration_complexity(listing)
        
        if complexity == 'Low':
            return '1-5 days'
        elif complexity == 'Medium':
            return '1-3 weeks'
        else:
            return '1-2 months'
    
    def check_reference_availability(self, listing):
        """Check reference availability"""
        publisher = self.get_publisher_name(listing).lower()
        major_vendors = ['oracle', 'microsoft', 'vmware', 'cisco', 'palo alto']
        
        if any(vendor in publisher for vendor in major_vendors):
            return 'Available'
        else:
            return 'Upon Request'
    
    def identify_competitors(self, listing):
        """Identify competitors"""
        category = self.get_category(listing).lower()
        
        competitor_map = {
            'security': 'Palo Alto, Fortinet, Check Point',
            'database': 'AWS RDS, Azure SQL, MongoDB',
            'analytics': 'Tableau, PowerBI, Qlik',
            'networking': 'Cisco, Juniper, Arista',
            'monitoring': 'Datadog, New Relic, Splunk'
        }
        
        for key, competitors in competitor_map.items():
            if key in category:
                return competitors
        
        return 'Market Leaders'
    
    def extract_value_proposition(self, listing):
        """Extract value proposition"""
        description = self.safe_get(listing, 'short-description', '')
        if description:
            # Get first sentence
            first_sentence = description.split('.')[0]
            return self.clean_text(first_sentence, 100)
        
        return 'Enterprise-grade cloud solution'
    
    def extract_key_features(self, listing):
        """Extract key features"""
        description = self.safe_get(listing, 'long-description', '') or self.safe_get(listing, 'short-description', '')
        
        if not description:
            return 'See product documentation'
        
        # Extract key phrases
        feature_indicators = ['provides', 'enables', 'includes', 'offers', 'supports']
        sentences = description.split('.')
        
        features = []
        for sentence in sentences[:2]:  # First 2 sentences
            if any(indicator in sentence.lower() for indicator in feature_indicators):
                clean_sentence = self.clean_text(sentence, 60)
                if clean_sentence:
                    features.append(clean_sentence)
        
        return '; '.join(features) if features else 'Advanced enterprise capabilities'
    
    def generate_marketplace_url(self, listing_id):
        """Generate marketplace URL"""
        clean_id = listing_id.replace('ocid1.marketplace.listing.', '')
        return f"https://cloudmarketplace.oracle.com/marketplace/en_US/listing/{clean_id}"
    
    def format_date(self, date_string):
        """Format date"""
        if not date_string:
            return ''
        
        try:
            if 'T' in str(date_string):
                date_obj = datetime.fromisoformat(str(date_string).replace('Z', '+00:00'))
                return date_obj.strftime('%Y-%m-%d')
        except:
            pass
        
        return str(date_string)[:10] if date_string else ''
    
    def determine_partner_level(self, listing):
        """Determine Oracle partner level"""
        publisher = self.safe_get(listing, 'publisher', {})
        publisher_name = self.get_publisher_name(listing).lower()
        
        if 'oracle' in publisher_name:
            return 'Oracle'
        elif self.safe_get(listing, 'oracle-validated'):
            return 'Oracle Validated'
        else:
            return 'ISV Partner'
    
    def generate_gov_sales_notes(self, listing, regions):
        """Generate government sales notes"""
        notes = []
        
        if regions.get('oc2_us_dod_east') or regions.get('oc2_us_dod_west'):
            notes.append('DoD Impact Level authorized')
        
        if regions.get('oc3_us_gov_east') or regions.get('oc3_us_gov_west'):
            notes.append('FedRAMP government ready')
        
        fedramp = self.determine_fedramp_status(listing)
        if fedramp != 'Not Specified':
            notes.append(fedramp)
        
        return '; '.join(notes) if notes else 'Commercial deployment only'
    
    def generate_sales_strategy_notes(self, listing, regions):
        """Generate sales strategy notes"""
        notes = []
        
        # Multi-realm availability
        realm_count = len([r for r in regions.keys() if regions[r]])
        if realm_count > 2:
            notes.append('Multi-realm solution - high flexibility')
        
        # Government focus
        if any(k.startswith('oc2_') or k.startswith('oc3_') for k in regions):
            notes.append('Government sales opportunity')
        
        # Pricing model
        pricing = self.get_pricing_model(listing)
        if pricing == 'Free':
            notes.append('Free tier available - easy customer adoption')
        elif pricing == 'Bring Your Own License':
            notes.append('BYOL model - leverage existing licenses')
        
        return '; '.join(notes) if notes else 'Standard commercial approach'
    
    def get_listing_text_content(self, listing):
        """Get all text content for analysis"""
        text_parts = [
            str(self.safe_get(listing, 'name', '')),
            str(self.safe_get(listing, 'short-description', '')),
            str(self.safe_get(listing, 'long-description', '')),
            str(self.safe_get(listing, 'tags', [])),
        ]
        return ' '.join(text_parts).lower()
    
    def is_security_focused(self, listing):
        """Check if security focused"""
        text = self.get_listing_text_content(listing)
        security_terms = ['security', 'firewall', 'vpn', 'encryption', 'auth', 'compliance']
        return any(term in text for term in security_terms)
    
    def calculate_compliance_score(self, row):
        """Calculate compliance score for government analysis"""
        score = 0
        
        if row['FedRAMP_Status'] != 'Not Specified':
            score += 3
        if row['DoD_Impact_Level'] != 'Not Specified':
            score += 3
        if row['CMMC_Level'] != 'Not Specified':
            score += 2
        if 'SOC 2' in str(row['Security_Certifications']):
            score += 1
        if 'ISO 27001' in str(row['Security_Certifications']):
            score += 1
        
        return min(10, score)
    
    def assess_gov_market_potential(self, row):
        """Assess government market potential"""
        if row['DoD_OC2'] == 'Yes' and row['Gov_Sales_Priority'] == 'CRITICAL':
            return 'Very High'
        elif row['US_Gov_OC3'] == 'Yes' and row['Gov_Sales_Priority'] in ['CRITICAL', 'HIGH']:
            return 'High'
        elif row['Gov_Sales_Priority'] == 'MEDIUM':
            return 'Medium'
        else:
            return 'Low'
    
    def classify_publisher_tier(self, publisher, stats):
        """Classify publisher tier"""
        if stats['total'] >= 10:
            return 'Tier 1 - Major Vendor'
        elif stats['total'] >= 5:
            return 'Tier 2 - Established'
        else:
            return 'Tier 3 - Emerging'
    
    def assess_publisher_strategic_value(self, stats):
        """Assess publisher strategic value"""
        gov_ratio = (stats['gov'] + stats['dod']) / max(stats['total'], 1)
        
        if gov_ratio >= 0.5 and stats['total'] >= 5:
            return 'High - Government Focus'
        elif stats['total'] >= 10:
            return 'High - Volume Partner'
        elif gov_ratio >= 0.3:
            return 'Medium - Gov Potential'
        else:
            return 'Standard'
    
    def recommend_publisher_action(self, publisher, stats):
        """Recommend action for publisher"""
        gov_products = stats['gov'] + stats['dod']
        
        if gov_products == 0 and stats['total'] >= 5:
            return 'Engage for government expansion'
        elif gov_products > 0 and stats['total'] >= 3:
            return 'Strategic partnership opportunity'
        else:
            return 'Monitor and maintain'

def main():
    """Main execution function"""
    print("🚀 Oracle Cloud Marketplace Global Sales Processor")
    print("Customer's Multi-Realm Configuration (OC1/OC2/OC3)")
    print("=" * 70)
    
    processor = OCIMarketplaceSalesProcessor()
    
    try:
        # Load data from all regions
        total_loaded = processor.load_all_regional_data()
        
        if total_loaded == 0:
            print("\n❌ No marketplace data found.")
            print("\n🔧 Troubleshooting:")
            print("   1. Run: ./extract_marketplace_customer.sh")
            print("   2. Verify OC1, OC2, OC3 profiles in ~/.oci/config")
            print("   3. Check marketplace_data/ directory exists")
            return
        
        # Consolidate unique listings
        unique_count = processor.consolidate_unique_listings()
        
        # Export to comprehensive Excel
        processor.export_to_excel('oracle_marketplace_global_sales_catalog.xlsx')
        
        print(f"\n🎉 Global Sales Catalog Complete!")
        print(f"📄 Excel File: oracle_marketplace_global_sales_catalog.xlsx")
        print(f"📊 Sheets: Complete Catalog | Government Opportunities | Publisher Analysis | Executive Summary")
        print(f"\n✅ Ready for global sales organization use!")
        
        print(f"\n📋 Next Steps:")
        print("   1. Distribute Excel file to regional sales teams")
        print("   2. Focus on 'Government Opportunities' sheet for federal sales")
        print("   3. Use 'Publisher Analysis' for partner strategies")
        print("   4. Reference 'Executive Summary' for leadership dashboards")
        
    except Exception as e:
        print(f"\n❌ Processing error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()